import openpyxl
from openpyxl.styles import Alignment
from prettytable import PrettyTable

# Cargar el archivo Excel existente
# En ruta_doc se ingresa la ruta específica del archivo
# en donde se encuentra el formulario a completar
# la ruta del archivo "ruta_doc se debe poner entre comillas simples ( ' ) o dobles ( " )
# y antes de la ruta se debe escribir una letra "r"m, esto es
# para poder definir una cadena cruda (raw string)
# sería de la siguiente manera ruta_doc = r'pegarrutadelarchivo'

ruta_doc = r'H:\Python_Proyects\UM\FORM_PRES_SALAS\PRESTAMO DE SALAS 2023 - PRUEBASPYTHON.xlsm'

workbook = openpyxl.load_workbook(ruta_doc, read_only=False, keep_vba=True)


# Seleccionar la hoja de trabajo que contiene la tabla
# En este caso el nombre de la hoja de la tabla debiese corresponder
# a todos los documentos debido a que se está interviniendo un formato

worksheet = workbook['REGISTRO DE PRESTAMO SALAS']

tabla_datos = []
for row in worksheet.iter_rows(min_row=7, max_row=258, values_only=True):
    tabla_datos.append(row)

# Imprimir los datos existentes en una tabla en la terminal,
# para poder visualizar los datos que contienen la tabla
# previo ingreso de información

tabla = PrettyTable()
tabla.field_names = ["N°", "FECHA", "NOMBRE COMPLETO", "CARRERA", "RUT RESPONSABLE","HORA INICIO","HORA TÉRMINO","AULA ASIGNADA","ACTIVIDAD","TOTAL ASISTENTES"]
for fila in tabla_datos:
    tabla.add_row(fila)
print("Datos existentes en la tabla:")
print('\n')
print(tabla)
print('\n')

# Obtener los valores de entrada del usuario

fecha = input("Ingresa la fecha: ")
nombre_completo = input("Ingresar nombre completo: ")
carrera = input("Ingresar carrera: ")
rut_responsable = input("Ingresar el RUT del/a responsable de la sala: ")
h_inicio = input("Ingresar hora de inicio: ")
h_termino = input("Ingresar hora de término: ")
aula_asignada = input("Ingresar sala facilitada: ")
actividad = input("Ingresar actividad: ")
n_asistentes = input("Ingresar N° de asistentes: ")

# Buscar la primera fila vacía en la columna A
# El rango es definido previamente y corresponde al
# tamaño de la tabla

fila_vacia = None
for i in range(7, 258):
    if worksheet.cell(row=i, column=2).value is None:
        fila_vacia = i
        break

# Si no se encontró una fila vacía, usar la siguiente fila después de la última fila llena
if fila_vacia is None:
    fila_vacia = worksheet.max_row + 1


# Escribir los valores en las celdas correspondientes
# Los valores son los ingresados por el usuario

worksheet.cell(row=fila_vacia, column=2, value=fecha).alignment = Alignment(horizontal='center', vertical='center')
worksheet.cell(row=fila_vacia, column=3, value=nombre_completo).alignment = Alignment(horizontal='center', vertical='center')
worksheet.cell(row=fila_vacia, column=4, value=carrera).alignment = Alignment(horizontal='center', vertical='center')
worksheet.cell(row=fila_vacia, column=5, value=rut_responsable).alignment = Alignment(horizontal='center', vertical='center')
worksheet.cell(row=fila_vacia, column=6, value=h_inicio).alignment = Alignment(horizontal='center', vertical='center')
worksheet.cell(row=fila_vacia, column=7, value=h_termino).alignment = Alignment(horizontal='center', vertical='center')
worksheet.cell(row=fila_vacia, column=8, value=aula_asignada).alignment = Alignment(horizontal='center', vertical='center')
worksheet.cell(row=fila_vacia, column=9, value=actividad).alignment = Alignment(horizontal='center', vertical='center')
worksheet.cell(row=fila_vacia, column=10, value=n_asistentes).alignment = Alignment(horizontal='center', vertical='center')


# Guardar los cambios en el archivo de Excel
workbook.save(ruta_doc)


#Imprime un mensaje indicandop que los cambios en el archivo Excel ya fueron guardados

print('\n')
print("Los datos se han guardado en el archivo de Excel")
